import openpyxl

def fetchUsername(SheetName):
#    wk = openpyxl.load_workbook("C:/Users/anand/PycharmProjects/RPA-Project/TestData/Credentials.xlsx")
    wk = openpyxl.load_workbook("TestData/Credentials.xlsx")
    sh = wk[SheetName]
    cell = sh.cell(2, 1)
    return cell.value

def fetchPwd(SheetName):
#   wk = openpyxl.load_workbook("C:/Users/anand/PycharmProjects/RPA-Project/TestData/Credentials.xlsx")
    wk = openpyxl.load_workbook("TestData/Credentials.xlsx")
    sh = wk[SheetName]
    cell = sh.cell(2, 2)
    return cell.value

